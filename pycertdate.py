# Required: pip install cryptography

import openpyxl
import ssl
from datetime import datetime, timezone
from cryptography import x509
from cryptography.hazmat.backends import default_backend
from datetime import datetime
import dns.resolver
import socket

# check if domain has name server entry
def has_ns_record(domain):
    try:
        dns.resolver.resolve(domain, 'NS') # in name service?
        return True
    except dns.resolver.NoAnswer:
        print(f"NS connection error.")
        return False
    except Exception as e:
        print(f"Error checking NS record for {domain}: {e}")
        return False

# get the certificate from the url
def get_certificate_expiration_date(url):
    try:
        cert = ssl.get_server_certificate((url, 443)) # fetch the certificate.
        cert_obj = x509.load_pem_x509_certificate(cert.encode(), default_backend()) # set certificate to object.
        expiration_date = cert_obj.not_valid_after_utc # set the certificate date.
        expiration_date = expiration_date.replace(tzinfo=timezone.utc).astimezone(tz=None)  # Convert to timezone-aware datetime object
        return expiration_date # return the certificate date.
    except Exception as e:
        print(f"Error fetching certificate expiration date for {url}: {e}")
        return str(e)  # Return the error message as a string

# update certificate date in excel file
def update_excel_with_certificate_expiration(file_path, url_column, py_date_column):
    try:
        wb = openpyxl.load_workbook(file_path) # open the excel
    except PermissionError as pe: # Tell the user the file might be open.
        print(f"PermissionError: Unable to open the file (already open?) {file_path}: {pe}")
        return

    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, min_col=url_column, max_col=url_column):
        url = row[0].value
        if not url:  # Skip processing if the URL is blank
            continue
        try:
            expiration_date = get_certificate_expiration_date(url) # try to get the date from certificate
            if isinstance(expiration_date, datetime): # good date, format it and set it
                # Convert datetime object to string in "mm/dd/yyyy" format for Excel compatibility
                expiration_date_str = expiration_date.strftime("%m/%d/%Y %H:%M") # format it for mm/dd/yyyy w/ military time.
                sheet.cell(row=row[0].row, column=py_date_column).value = expiration_date_str # update the column.
                now = datetime.now() # set right now date/time
                now = now.strftime("%m/%d/%Y %H:%M") # format it for mm/dd/yyyy w/ military time.
                sheet.cell(row=row[0].row, column=py_last_check).value = now # update the column.
                print(f"Updated date for URL: {url}") # tell the user on console.
            else: # not a date, do other things.
                sheet.cell(row=row[0].row, column=py_date_column).value = expiration_date 
        except socket.gaierror: # can't connect to it.
            sheet.cell(row=row[0].row, column=py_date_column).value = "Contact Owner - Invalid"
        except Exception as e: # all other errors.
            print(f"Error processing {url}: {e}")
            sheet.cell(row=row[0].row, column=py_date_column).value = str(e)  # Put the error text in the column

    wb.save(file_path) # save the excel file.
    wb.close()  # Close the workbook to release the file handle.

if __name__ == "__main__": # settings
    file_path = "pycertdate.xlsx"  # Change this to the path of your Excel file
    url_column = 1  # Column containing URL
    py_date_column = 3  # Column where you want to update the expiration date
    py_last_check = 4  # Column where you want to update the last check date
    update_excel_with_certificate_expiration(file_path, url_column, py_date_column) # write it to excel.
